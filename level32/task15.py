import pandas as pd
from openpyxl import load_workbook
import shutil

# Definimos os caminhos para os arquivos CSV e Excel
csv_file_path = 'employees.csv'  # Caminho para o arquivo CSV
excel_file_path = 'employees_data.xlsx'  # Caminho para o arquivo Excel existente

# Leitura dos dados do arquivo CSV
csv_data = pd.read_csv(csv_file_path)

# Criação de cópia de segurança do arquivo Excel
backup_path = excel_file_path.replace('.xlsx', '_backup.xlsx')
shutil.copy(excel_file_path, backup_path)

# Carregamento do arquivo Excel existente
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Definição da linha inicial para adicionar novos dados
start_row = sheet.max_row + 1

# Adição dos dados do CSV ao Excel
for index, row in csv_data.iterrows():
    for col_num, value in enumerate(row, start=1):
        # Escrevemos o valor na célula correspondente
        sheet.cell(row=start_row + index, column=col_num, value=value)

# Salvamos as alterações no arquivo Excel
workbook.save(excel_file_path)